"""
LinkedIn Talent Discovery Endpoints — powered by Unipile

All Unipile HTTP calls are made server-side only.
API keys are read from environment; never exposed to the browser.
"""
from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field

from app.api.v2.deps import require_user
from app.auth.clerk import ClerkUser
from app.config import get_settings
from app.services.ai.factory import get_ai
from app.services.db.supabase_service import get_db_admin_service
from app.services.email_service import EmailService
from app.utils.responses import api_error, ok

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/linkedin")

# ─── Excel Profile Pool ───────────────────────────────────────────────────────

_EXCEL_DIR = Path(__file__).resolve().parents[5] / "data"
_EXCEL_PATH = _EXCEL_DIR / "linkedin_profiles_pool.xlsx"

_EXCEL_HEADERS = [
    "Timestamp", "Job ID", "Job Title", "Name", "Headline", "Location",
    "Current Role", "Current Company", "Skills", "LinkedIn URL",
    "Match Score", "ATS Label",
]


def _ats_label(score: int) -> str:
    if score >= 80: return "Excellent"
    if score >= 60: return "Strong"
    if score >= 40: return "Fair"
    return "Weak"


def _append_profiles_to_excel(job_id: str, job_title: str, profiles: List[Dict[str, Any]]) -> None:
    """Append retrieved LinkedIn profiles to the running Excel data pool.

    Creates the file and header row on first use. De-duplicates by LinkedIn URL
    or provider_id so repeated searches don't add duplicate rows.
    """
    try:
        import openpyxl  # lazy import — only needed here
        from openpyxl.styles import Font, PatternFill, Alignment

        _EXCEL_DIR.mkdir(parents=True, exist_ok=True)

        # Load or create workbook
        if _EXCEL_PATH.exists():
            wb = openpyxl.load_workbook(_EXCEL_PATH)
            ws = wb.active
            # Collect existing LinkedIn URLs + provider IDs to de-duplicate
            existing_keys: set = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                url = row[9] if row and len(row) > 9 else None  # LinkedIn URL column
                if url:
                    existing_keys.add(str(url).strip())
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "LinkedIn Profile Pool"
            existing_keys: set = set()
            # Write styled header row
            header_row = ws.append  # use below
            ws.append(_EXCEL_HEADERS)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0077B5", end_color="0077B5", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            # Set column widths
            col_widths = [20, 36, 25, 25, 40, 20, 30, 25, 60, 50, 12, 12]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        added = 0
        for p in profiles:
            public_id = p.get("public_identifier") or p.get("provider_id") or ""
            linkedin_url = (
                p.get("public_profile_url")
                or (f"https://www.linkedin.com/in/{public_id}" if public_id else "")
            ).strip()

            # De-duplicate
            dedup_key = linkedin_url or public_id
            if dedup_key and dedup_key in existing_keys:
                continue
            if dedup_key:
                existing_keys.add(dedup_key)

            work = p.get("work_experience") or []
            current = work[0] if work else {}
            skills_raw = p.get("skills") or []
            skills_str = ", ".join(
                s.get("name", str(s)) if isinstance(s, dict) else str(s)
                for s in skills_raw[:15]
            )
            score = int(p.get("match_score") or 0)
            name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip()

            ws.append([
                ts,
                job_id,
                job_title,
                name,
                p.get("headline", ""),
                p.get("location", ""),
                current.get("position", ""),
                current.get("company", ""),
                skills_str,
                linkedin_url,
                score,
                _ats_label(score) if score > 0 else "",
            ])
            added += 1

        if added > 0:
            wb.save(_EXCEL_PATH)
            logger.info("[excel_pool] Added %d new profiles to %s", added, _EXCEL_PATH)
        else:
            logger.info("[excel_pool] No new profiles to add (all duplicates).")

    except Exception as exc:
        # Non-fatal — log but never break the search response
        logger.warning("[excel_pool] Failed to update Excel pool: %s", exc)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _unipile_headers() -> Dict[str, str]:
    settings = get_settings()
    return {
        "X-API-KEY": settings.unipile_api_key,
        "accept": "application/json",
        "content-type": "application/json",
    }


def _unipile_base() -> str:
    settings = get_settings()
    dsn = (settings.unipile_dsn or "").rstrip("/")
    return f"{dsn}/api/v1"


def _check_unipile_configured() -> bool:
    s = get_settings()
    return bool(s.unipile_api_key and s.unipile_dsn)


_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _extract_email_from_profile(profile: Dict[str, Any]) -> Optional[str]:
    """Try to extract a real email address from a Unipile profile dict.

    Unipile may expose email in several locations depending on API version:
      - profile["email"]          – direct field on open / premium profiles
      - profile["emails"][0]      – list of email objects {"address": "..."}
      - profile["contact_info"]["emails"][0]["address"]
      - profile["member_info"]["email"]
    Returns None if no valid email is found.
    """
    # 1. Direct top-level field
    direct = str(profile.get("email") or "").strip()
    if direct and _EMAIL_RE.match(direct):
        return direct

    # 2. Emails list: [{"address": "...", ...}] or ["email@example.com", ...]
    for entry in (profile.get("emails") or []):
        if isinstance(entry, dict):
            addr = str(entry.get("address") or "").strip()
        else:
            addr = str(entry).strip()
        if addr and _EMAIL_RE.match(addr):
            return addr

    # 3. contact_info.emails
    contact = profile.get("contact_info") or {}
    for entry in (contact.get("emails") or []):
        if isinstance(entry, dict):
            addr = str(entry.get("address") or "").strip()
        else:
            addr = str(entry).strip()
        if addr and _EMAIL_RE.match(addr):
            return addr

    # 4. member_info.email
    member = profile.get("member_info") or {}
    addr = str(member.get("email") or "").strip()
    if addr and _EMAIL_RE.match(addr):
        return addr

    return None


# ─── Schemas ──────────────────────────────────────────────────────────────────

class GenerateFiltersRequest(BaseModel):
    job_id: str
    title: Optional[str] = None
    description: Optional[str] = None
    must_have_skills: Optional[List[str]] = None
    good_to_have_skills: Optional[List[str]] = None
    min_experience_years: Optional[int] = None
    location: Optional[str] = None


class GenerateFiltersResponse(BaseModel):
    keywords: str = ""
    skills: List[str] = []
    job_titles: List[str] = []
    similar_roles: List[str] = []
    experience_min: int = 0
    experience_max: int = 10
    industry: str = ""
    location: str = ""
    seniority: str = ""
    preferred_education: str = ""


class LinkedInSearchRequest(BaseModel):
    job_id: str
    filters: Dict[str, Any] = {}
    candidate_count: int = Field(default=10, ge=1, le=30)
    title: Optional[str] = None
    description: Optional[str] = None
    must_have_skills: Optional[List[str]] = None


class CandidateProfile(BaseModel):
    provider_id: Optional[str] = None
    public_identifier: Optional[str] = None
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    headline: Optional[str] = None
    location: Optional[str] = None
    profile_picture_url: Optional[str] = None
    summary: Optional[str] = None
    work_experience: List[Dict[str, Any]] = []
    education: List[Dict[str, Any]] = []
    skills: List[Any] = []
    languages: List[Any] = []
    certifications: List[Any] = []
    follower_count: Optional[int] = None
    connections_count: Optional[int] = None
    is_open_profile: Optional[bool] = None
    websites: List[str] = []
    match_score: int = 0
    ai_summary: str = ""
    public_profile_url: Optional[str] = None


class RankCandidatesRequest(BaseModel):
    job_id: str
    title: str
    description: str
    must_have_skills: List[str] = []
    good_to_have_skills: List[str] = []
    min_experience_years: Optional[int] = None
    candidates: List[Dict[str, Any]]


class GenerateEmailRequest(BaseModel):
    candidate: Dict[str, Any]
    job_title: str
    company_name: str
    job_description: Optional[str] = None
    recruiter_name: Optional[str] = None
    job_apply_url: Optional[str] = None


class SaveCandidateRequest(BaseModel):
    job_id: str
    search_id: Optional[str] = None
    profile: Dict[str, Any]
    match_score: int = 0
    ai_summary: Optional[str] = None


class UpdateCandidateRequest(BaseModel):
    notes: Optional[str] = None
    tags: Optional[List[str]] = None
    status: Optional[str] = None


class SendMessageRequest(BaseModel):
    account_id: str
    provider_id: str
    message: str


# ─── Endpoints ────────────────────────────────────────────────────────────────


@router.get("/accounts")
async def get_linkedin_accounts(user: ClerkUser = Depends(require_user)):
    """
    Check whether the recruiter has a LinkedIn account connected via Unipile.
    Returns list of connected accounts.
    """
    if not _check_unipile_configured():
        return ok({
            "connected": False,
            "accounts": [],
            "message": "Unipile is not configured. Set UNIPILE_DSN and UNIPILE_API_KEY in the backend .env."
        })

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                f"{_unipile_base()}/accounts",
                headers=_unipile_headers(),
                params={"limit": 20},
            )
            if resp.status_code != 200:
                return ok({"connected": False, "accounts": [], "message": "Failed to reach Unipile"})

            data = resp.json()
            accounts = data.get("items", [])
            linkedin_accounts = [a for a in accounts if a.get("type") == "LINKEDIN" or a.get("provider") == "LINKEDIN"]

            return ok({
                "connected": len(linkedin_accounts) > 0,
                "accounts": linkedin_accounts,
            })
    except Exception as e:
        return ok({"connected": False, "accounts": [], "message": str(e)})


@router.post("/generate-filters")
async def generate_filters(
    payload: GenerateFiltersRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    AI-analyze the Job Description and return pre-populated LinkedIn search filters.
    """
    description = (payload.description or "").strip()
    title = (payload.title or "").strip()
    skills_str = ", ".join(payload.must_have_skills or [])
    experience = payload.min_experience_years or 0
    location = payload.location or ""

    prompt = f"""You are an expert technical recruiter. Analyze this job description and extract LinkedIn search parameters.

Job Title: {title}
Skills Required: {skills_str}
Min Experience: {experience} years
Location: {location}

Job Description:
{description[:4000]}

Return ONLY valid JSON in this exact format:
{{
  "keywords": "primary search keyword phrase (3-5 words max)",
  "skills": ["skill1", "skill2", "skill3", "skill4", "skill5"],
  "job_titles": ["Exact Job Title", "Alternative Title 1", "Alternative Title 2"],
  "similar_roles": ["Similar Role 1", "Similar Role 2"],
  "experience_min": {experience},
  "experience_max": {max(experience + 5, 10)},
  "industry": "Primary industry (e.g. Technology, Finance)",
  "location": "{location}",
  "seniority": "Entry/Mid/Senior/Lead/Manager/Director/VP",
  "preferred_education": "Bachelor's/Master's/PhD or empty string"
}}

Rules:
- keywords: the best 3-5 word phrase to find these candidates on LinkedIn
- skills: pick the top 5 most searchable skills from the JD
- job_titles: include the exact title AND 2 common variations
- seniority: infer from experience years and title
"""

    try:
        ai = get_ai()
        result = await ai.generate_json(prompt, temperature=0.1, max_tokens=800, timeout_s=25)

        if not isinstance(result, dict):
            raise ValueError("Invalid AI response")

        return ok(GenerateFiltersResponse(
            keywords=result.get("keywords", title),
            skills=result.get("skills", payload.must_have_skills or []),
            job_titles=result.get("job_titles", [title]),
            similar_roles=result.get("similar_roles", []),
            experience_min=result.get("experience_min", experience),
            experience_max=result.get("experience_max", experience + 5),
            industry=result.get("industry", ""),
            location=result.get("location", location),
            seniority=result.get("seniority", ""),
            preferred_education=result.get("preferred_education", ""),
        ).model_dump())

    except Exception as e:
        return api_error(message=f"Failed to generate filters: {e}", status_code=502)


@router.post("/search")
async def search_linkedin_candidates(
    payload: LinkedInSearchRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    Search LinkedIn for candidates via Unipile.
    Uses Unipile's LinkedIn People Search API.
    Returns up to `candidate_count` profiles (max 30).
    """
    settings = get_settings()

    # Enforce cap
    count = min(max(payload.candidate_count, 1), 30)
    filters = payload.filters or {}

    if not _check_unipile_configured():
        return api_error(
            message="Unipile is not configured. Please add UNIPILE_DSN and UNIPILE_API_KEY to the backend .env.",
            status_code=503
        )

    account_id = settings.unipile_account_id

    # Build Unipile search query from filters
    keywords = filters.get("keywords", "")
    location = filters.get("location", "")
    job_titles = filters.get("job_titles", [])
    industry = filters.get("industry", "")

    if not keywords and job_titles:
        keywords = job_titles[0]
    if not keywords and payload.title:
        keywords = payload.title

    profiles: List[Dict[str, Any]] = []

    try:
        async with httpx.AsyncClient(timeout=90) as client:
            # ── Unipile LinkedIn Search API ───────────────────────────────────
            # POST /api/v1/linkedin/search
            #   Query param: account_id, limit
            #   JSON body:   { api: "classic", category: "people", keywords }
            #
            # CONFIRMED from Unipile 400 error schema:
            #   field names are "api" and "category" (not "service"/"type")
            #
            # IMPORTANT: Unipile's location filter requires a LinkedIn location ID,
            # NOT a plain text string. Passing a text string causes a 400.
            # We fold location into the keywords string so searches still work
            # geographically without triggering parameter validation errors.
            search_body: Dict[str, Any] = {
                "api": "classic",
                "category": "people",
            }
            # Combine keywords + location into a single keywords string.
            # Unipile's location filter requires a LinkedIn location ID (not plain text).
            # Folding location into keywords is the safe approach.
            kw_parts = [p for p in [keywords, location] if p]
            combined_keywords = " ".join(kw_parts).strip()
            if combined_keywords:
                search_body["keywords"] = combined_keywords

            fetch_limit = min(count + 5, 49)  # overfetch slightly; trim after

            search_resp = await client.post(
                f"{_unipile_base()}/linkedin/search",
                headers=_unipile_headers(),
                params={"account_id": account_id, "limit": fetch_limit},
                json=search_body,
            )

            if search_resp.status_code != 200:
                return api_error(
                    message=f"LinkedIn search failed (status {search_resp.status_code}): {search_resp.text[:500]}",
                    status_code=502,
                )

            search_data = search_resp.json()
            raw_results = search_data.get("items", [])[:count]

            # ── Enrich results concurrently ────────────────────────────────────
            # Sequential fetches were taking ~8-9s per profile = 90s for 10 profiles.
            # asyncio.gather runs all fetches in parallel → total time ≈ slowest single fetch.
            import asyncio

            async def _enrich_one(raw: Dict[str, Any]) -> Dict[str, Any]:
                identifier = (
                    raw.get("public_identifier")
                    or raw.get("provider_id")
                    or raw.get("id")
                )
                if not identifier:
                    return raw
                try:
                    profile_resp = await client.get(
                        f"{_unipile_base()}/users/{identifier}",
                        headers=_unipile_headers(),
                        params={
                            "account_id": account_id,
                            "linkedin_sections": "*",
                        },
                    )
                    if profile_resp.status_code == 200:
                        full = profile_resp.json()
                        # Merge search-result fields that may not be in profile
                        full.setdefault("headline", raw.get("headline", ""))
                        full.setdefault("location", raw.get("location", ""))
                        full.setdefault("public_profile_url", raw.get("public_profile_url", ""))
                        # Extract email if Unipile returns one (open / premium profiles)
                        email = _extract_email_from_profile(full)
                        if email:
                            full["email"] = email
                        return full
                    else:
                        # Fallback to search-result data with normalized keys.
                        # Docs confirm the field is "provider_id" in Unipile responses.
                        return {
                            "provider_id": raw.get("provider_id") or raw.get("id"),
                            "public_identifier": raw.get("public_identifier"),
                            "first_name": (raw.get("name") or "").split(" ")[0],
                            "last_name": " ".join((raw.get("name") or "").split(" ")[1:]),
                            "headline": raw.get("headline", ""),
                            "location": raw.get("location", ""),
                            "profile_picture_url": raw.get("profile_picture_url"),
                            "profile_picture_url_large": raw.get("profile_picture_url_large"),
                            "public_profile_url": raw.get("public_profile_url"),
                            "network_distance": raw.get("network_distance"),
                            "is_open_profile": raw.get("is_open_profile"),
                            "is_premium": raw.get("is_premium"),
                            "is_relationship": raw.get("is_relationship"),
                        }
                except Exception:
                    return raw

            enriched: List[Dict[str, Any]] = list(
                await asyncio.gather(*[_enrich_one(r) for r in raw_results])
            )
            profiles = enriched

    except httpx.TimeoutException:
        return api_error(message="LinkedIn search timed out. Please try again.", status_code=504)
    except Exception as e:
        return api_error(message=f"LinkedIn search failed: {e}", status_code=502)


    # Save search record to Supabase
    db = get_db_admin_service()
    try:
        search_record = db.client.table("linkedin_searches").insert({
            "job_id": payload.job_id,
            "recruiter_id": user.id,
            "filters": filters,
            "candidate_count_requested": count,
            "profiles_retrieved": len(profiles),
        }).execute()
        search_id = search_record.data[0]["id"] if search_record.data else None
    except Exception:
        search_id = None

    # ── Auto-save profiles to Excel data pool ─────────────────────────────────
    _append_profiles_to_excel(
        job_id=payload.job_id,
        job_title=payload.title or "",
        profiles=profiles,
    )

    return ok({
        "search_id": search_id,
        "requested": count,
        "retrieved": len(profiles),
        "profiles": profiles,
    })


@router.post("/profile")
async def get_full_profile(
    identifier: str = Query(..., description="LinkedIn public identifier or provider_id"),
    account_id: Optional[str] = Query(None),
    user: ClerkUser = Depends(require_user),
):
    """
    Fetch full LinkedIn profile via Unipile.
    Uses GET /users/{identifier}?linkedin_sections=*
    """
    if not _check_unipile_configured():
        return api_error(message="Unipile is not configured.", status_code=503)

    settings = get_settings()
    acct = account_id or settings.unipile_account_id

    try:
        async with httpx.AsyncClient(timeout=20) as client:
            resp = await client.get(
                f"{_unipile_base()}/users/{identifier}",
                headers=_unipile_headers(),
                params={
                    "account_id": acct,
                    "linkedin_sections": "*",
                },
            )
            if resp.status_code == 200:
                return ok(resp.json())
            else:
                return api_error(
                    message=f"Failed to fetch profile: {resp.status_code}",
                    status_code=resp.status_code
                )
    except httpx.TimeoutException:
        return api_error(message="Profile fetch timed out.", status_code=504)
    except Exception as e:
        return api_error(message=f"Profile fetch failed: {e}", status_code=502)


@router.post("/rank-candidates")
async def rank_candidates(
    payload: RankCandidatesRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    AI-rank a list of LinkedIn profiles against the Job Description.
    Returns profiles sorted by match_score desc with AI summary per candidate.
    """
    if not payload.candidates:
        return ok({"ranked": []})

    must_have = ", ".join(payload.must_have_skills) if payload.must_have_skills else "Not specified"
    good_to_have = ", ".join(payload.good_to_have_skills) if payload.good_to_have_skills else "None"

    # Build compact candidate summaries for the prompt
    candidate_summaries = []
    for i, c in enumerate(payload.candidates):
        work = c.get("work_experience", [])
        current = work[0] if work else {}
        edu = c.get("education", [])
        first_edu = edu[0] if edu else {}
        skills_raw = c.get("skills", [])
        skills_list = [s.get("name", s) if isinstance(s, dict) else str(s) for s in skills_raw[:10]]

        candidate_summaries.append({
            "index": i,
            "name": f"{c.get('first_name','')} {c.get('last_name','')}".strip(),
            "headline": c.get("headline", ""),
            "location": c.get("location", ""),
            "current_role": current.get("position", ""),
            "current_company": current.get("company", ""),
            "years_experience": len(work),
            "skills": skills_list,
            "education": first_edu.get("school", ""),
            "degree": first_edu.get("degree", ""),
        })

    prompt = f"""You are an expert technical recruiter. Score and rank each candidate against this job.

JOB: {payload.title}
MUST-HAVE SKILLS: {must_have}
NICE-TO-HAVE: {good_to_have}
MIN EXPERIENCE: {payload.min_experience_years or 0} years

CANDIDATES:
{json.dumps(candidate_summaries, indent=2)}

For EACH candidate, return a JSON object with:
- "index": (same integer as input)
- "match_score": integer 0-100 based on skills match, experience, title relevance
- "ai_summary": 2-3 sentence explanation of why this candidate is or isn't a fit
- "strengths": list of 2-3 matching points
- "gaps": list of 1-2 potential gaps (or empty list if strong match)

Return a JSON array of these objects, sorted by match_score descending.
Return ONLY the JSON array, no markdown, no extra text."""

    try:
        ai = get_ai()
        result = await ai.generate_json(prompt, temperature=0.1, max_tokens=3000, timeout_s=40)

        if not isinstance(result, list):
            result = []

        # Merge scores back into original profiles
        score_map = {item["index"]: item for item in result if isinstance(item, dict)}
        ranked = []
        for i, profile in enumerate(payload.candidates):
            scoring = score_map.get(i, {})
            profile_copy = dict(profile)
            profile_copy["match_score"] = scoring.get("match_score", 0)
            profile_copy["ai_summary"] = scoring.get("ai_summary", "")
            profile_copy["strengths"] = scoring.get("strengths", [])
            profile_copy["gaps"] = scoring.get("gaps", [])
            ranked.append(profile_copy)

        ranked.sort(key=lambda x: x.get("match_score", 0), reverse=True)
        return ok({"ranked": ranked})

    except Exception as e:
        # Return profiles unranked if AI fails
        return ok({"ranked": payload.candidates, "error": str(e)})


@router.post("/generate-email")
async def generate_outreach_email(
    payload: GenerateEmailRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    Generate a personalized outreach email for a LinkedIn candidate.
    """
    candidate = payload.candidate
    name = f"{candidate.get('first_name', '')} {candidate.get('last_name', '')}".strip() or "there"
    first_name = candidate.get("first_name", "there")
    headline = candidate.get("headline", "")
    work = candidate.get("work_experience", [])
    current_role = work[0].get("position", "") if work else ""
    current_company = work[0].get("company", "") if work else ""
    skills_raw = candidate.get("skills", [])
    skills_list = [s.get("name", s) if isinstance(s, dict) else str(s) for s in skills_raw[:5]]
    skills_str = ", ".join(skills_list) if skills_list else "your technical background"

    recruiter_name = payload.recruiter_name or "The Hiring Team"

    apply_url = (payload.job_apply_url or "").strip()
    apply_instruction = (
        f"\n7. Include a direct apply link at the end: {apply_url}"
        if apply_url else ""
    )
    apply_note = (
        f"\n- Apply Link: {apply_url} (include this in the email so the candidate can apply directly)"
        if apply_url else ""
    )

    prompt = f"""Write a professional, personalized LinkedIn/email outreach message to recruit this candidate.

CANDIDATE:
- Name: {name}
- Current Role: {current_role} at {current_company}
- Headline: {headline}
- Key Skills: {skills_str}

OPPORTUNITY:
- Job Title: {payload.job_title}
- Company: {payload.company_name}
- Description excerpt: {(payload.job_description or '')[:500]}{apply_note}

RECRUITER: {recruiter_name}

Write a professional outreach email. Structure:
1. Personalized greeting using first name
2. Why YOU noticed THEM specifically (reference their specific background)
3. Brief company introduction (1 sentence)
4. The opportunity (2-3 sentences)
5. Clear call-to-action (suggest a 15-min call){apply_instruction}
6. Professional sign-off

Tone: Professional but warm. NOT generic. NOT salesy. Under 220 words.

Return JSON with:
{{
  "subject": "Email subject line",
  "body": "Full email body (plain text, use \\n for line breaks)"
}}"""

    try:
        ai = get_ai()
        result = await ai.generate_json(prompt, temperature=0.7, max_tokens=600, timeout_s=20)

        if not isinstance(result, dict):
            raise ValueError("Invalid AI response")

        return ok({
            "subject": result.get("subject", f"Opportunity for {payload.job_title} at {payload.company_name}"),
            "body": result.get("body", ""),
        })

    except Exception as e:
        # Fallback template
        fallback_subject = f"Opportunity for {payload.job_title} at {payload.company_name}"
        fallback_body = (
            f"Hi {first_name},\n\n"
            f"I came across your profile and was impressed by your experience as {current_role or 'a professional in your field'}.\n\n"
            f"I'm reaching out about an exciting {payload.job_title} opportunity at {payload.company_name}. "
            f"Based on your background, I believe you could be a great fit.\n\n"
            f"Would you be open to a quick 15-minute call to learn more?\n\n"
            f"Best regards,\n{recruiter_name}"
        )
        return ok({"subject": fallback_subject, "body": fallback_body})


@router.post("/save-candidate")
async def save_candidate(
    payload: SaveCandidateRequest,
    user: ClerkUser = Depends(require_user),
):
    """Save a LinkedIn candidate profile to the database."""
    db = get_db_admin_service()
    profile = payload.profile

    try:
        existing = db.client.table("linkedin_saved_candidates").select("id").eq(
            "job_id", payload.job_id
        ).eq(
            "recruiter_id", user.id
        ).eq(
            "linkedin_id", profile.get("provider_id", "")
        ).execute()

        if existing.data:
            # Update existing
            result = db.client.table("linkedin_saved_candidates").update({
                "profile_data": profile,
                "match_score": payload.match_score,
                "ai_summary": payload.ai_summary or "",
            }).eq("id", existing.data[0]["id"]).execute()
            return ok({"id": existing.data[0]["id"], "updated": True})

        # Insert new
        result = db.client.table("linkedin_saved_candidates").insert({
            "job_id": payload.job_id,
            "recruiter_id": user.id,
            "search_id": payload.search_id,
            "linkedin_id": profile.get("provider_id"),
            "public_identifier": profile.get("public_identifier"),
            "profile_data": profile,
            "match_score": payload.match_score,
            "ai_summary": payload.ai_summary or "",
            "status": "saved",
        }).execute()

        return ok({"id": result.data[0]["id"] if result.data else None, "created": True})

    except Exception as e:
        err_str = str(e)
        # Graceful degradation: if the table doesn't exist yet (migration not run),
        # return a mock success so the UI toast doesn't show an error.
        # The candidate is still displayed as "saved" locally in the UI.
        if "relation" in err_str and "does not exist" in err_str:
            return ok({"id": None, "created": True, "note": "Table not yet migrated — run linkedin_talent.sql in Supabase."})
        return api_error(message=f"Failed to save candidate: {e}", status_code=500)


@router.patch("/candidate/{candidate_id}")
async def update_candidate(
    candidate_id: str,
    payload: UpdateCandidateRequest,
    user: ClerkUser = Depends(require_user),
):
    """Update notes, tags, or status on a saved candidate."""
    db = get_db_admin_service()

    update_data: Dict[str, Any] = {}
    if payload.notes is not None:
        update_data["notes"] = payload.notes
    if payload.tags is not None:
        update_data["tags"] = payload.tags
    if payload.status is not None:
        update_data["status"] = payload.status
        if payload.status == "contacted":
            update_data["contacted_at"] = "now()"

    if not update_data:
        return ok({"updated": False, "message": "Nothing to update"})

    try:
        result = db.client.table("linkedin_saved_candidates").update(update_data).eq(
            "id", candidate_id
        ).eq("recruiter_id", user.id).execute()
        return ok({"updated": True})
    except Exception as e:
        return api_error(message=f"Failed to update candidate: {e}", status_code=500)


@router.get("/saved-candidates/{job_id}")
async def get_saved_candidates(
    job_id: str,
    user: ClerkUser = Depends(require_user),
):
    """Get all saved LinkedIn candidates for a job."""
    db = get_db_admin_service()
    try:
        result = db.client.table("linkedin_saved_candidates").select("*").eq(
            "job_id", job_id
        ).eq("recruiter_id", user.id).order("match_score", desc=True).execute()
        return ok({"candidates": result.data or []})
    except Exception as e:
        # Always degrade gracefully — table may not be migrated yet.
        # Log the real error server-side; return empty list to the client.
        import logging
        logging.getLogger(__name__).warning("get_saved_candidates: %s", e)
        return ok({"candidates": []})


@router.delete("/saved-candidates/{job_id}/{linkedin_id}")
async def unsave_candidate(
    job_id: str,
    linkedin_id: str,
    user: ClerkUser = Depends(require_user),
):
    """Remove a saved LinkedIn candidate profile."""
    db = get_db_admin_service()
    try:
        db.client.table("linkedin_saved_candidates").delete().eq(
            "job_id", job_id
        ).eq("recruiter_id", user.id).eq("linkedin_id", linkedin_id).execute()
        return ok({"deleted": True})
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("unsave_candidate: %s", e)
        return ok({"deleted": False})


@router.get("/search-history/{job_id}")
async def get_search_history(
    job_id: str,
    user: ClerkUser = Depends(require_user),
):
    """Get previous LinkedIn searches for a job."""
    db = get_db_admin_service()
    try:
        result = db.client.table("linkedin_searches").select("*").eq(
            "job_id", job_id
        ).eq("recruiter_id", user.id).order("created_at", desc=True).limit(20).execute()
        return ok({"searches": result.data or []})
    except Exception as e:
        # Always degrade gracefully — table may not be migrated yet.
        import logging
        logging.getLogger(__name__).warning("get_search_history: %s", e)
        return ok({"searches": []})


@router.post("/send-message")
async def send_linkedin_message(
    payload: SendMessageRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    Send a LinkedIn message via Unipile.
    Creates or finds a chat with the target user and sends a message.
    """
    if not _check_unipile_configured():
        return api_error(message="Unipile is not configured.", status_code=503)

    try:
        async with httpx.AsyncClient(timeout=20) as client:
            # Per Unipile docs: POST /chats creates or retrieves an existing chat
            # with the attendee. This works for 1st-degree connections.
            # For non-connections, LinkedIn requires sending an invitation first;
            # POST /chats will fail if no prior relationship exists.
            chat_resp = await client.post(
                f"{_unipile_base()}/chats",
                headers=_unipile_headers(),
                json={
                    "account_id": payload.account_id,
                    "attendees_ids": [payload.provider_id],
                }
            )

            if chat_resp.status_code not in (200, 201):
                err_text = chat_resp.text[:300]
                # Provide an actionable message if likely a non-connection
                hint = (
                    " This candidate may not be a 1st-degree connection. "
                    "Try sending a LinkedIn invitation first."
                    if chat_resp.status_code in (400, 403)
                    else ""
                )
                return api_error(
                    message=f"Failed to create LinkedIn chat ({chat_resp.status_code}): {err_text}{hint}",
                    status_code=502
                )

            chat_data = chat_resp.json()
            chat_id = chat_data.get("id") or chat_data.get("object_id")

            if not chat_id:
                return api_error(message="Could not retrieve chat ID from Unipile.", status_code=502)

            # Send the message
            msg_resp = await client.post(
                f"{_unipile_base()}/chats/{chat_id}/messages",
                headers=_unipile_headers(),
                json={
                    "account_id": payload.account_id,
                    "text": payload.message,
                }
            )

            if msg_resp.status_code in (200, 201):
                return ok({"sent": True, "chat_id": chat_id})
            else:
                return api_error(
                    message=f"Failed to send message: {msg_resp.status_code} — {msg_resp.text[:300]}",
                    status_code=502
                )

    except httpx.TimeoutException:
        return api_error(message="Message send timed out.", status_code=504)
    except Exception as e:
        return api_error(message=f"Failed to send LinkedIn message: {e}", status_code=502)


# ─── Outreach Email via SMTP ───────────────────────────────────────────────────

class SendOutreachEmailRequest(BaseModel):
    to_email: str = Field(..., description="Candidate's email address")
    subject: str
    body: str
    candidate_name: Optional[str] = None
    job_title: Optional[str] = None
    job_apply_url: Optional[str] = None


@router.post("/send-outreach-email")
async def send_outreach_email(
    payload: SendOutreachEmailRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    Send a recruitment outreach email directly via SMTP (Hostinger).
    Uses the existing EmailService — no external provider required.
    """
    to = payload.to_email.strip()
    if not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", to):
        return api_error(message="Invalid recipient email address.", status_code=400)

    # Build HTML body from plain text
    html_body = payload.body.replace("\n", "<br>").replace("  ", "&nbsp;&nbsp;")

    apply_url = (payload.job_apply_url or "").strip()
    apply_section = ""
    if apply_url and re.match(r"https?://", apply_url):
        apply_section = f"""
        <div style="margin-top: 28px; text-align: center;">
            <a href="{apply_url}"
               style="display: inline-block; padding: 14px 32px; background: #0077B5;
                      color: #ffffff; text-decoration: none; border-radius: 8px;
                      font-family: Arial, sans-serif; font-size: 15px; font-weight: 600;
                      letter-spacing: 0.3px;">
                Apply for this Position &rarr;
            </a>
            <p style="margin-top: 10px; font-size: 12px; color: #9ca3af;">
                Or copy this link: <a href="{apply_url}" style="color: #0077B5;">{apply_url}</a>
            </p>
        </div>"""

    html_content = f"""
    <div style="font-family: Arial, sans-serif; max-width: 680px; margin: 0 auto;
                padding: 32px 24px; color: #1a1a2e;">
        <div style="border-bottom: 2px solid #0077B5; padding-bottom: 16px; margin-bottom: 24px;">
            <span style="font-size: 13px; color: #0077B5; font-weight: 600;">Rekshift · LinkedIn Outreach</span>
        </div>
        <div style="font-size: 15px; line-height: 1.7; color: #333;">
            {html_body}
        </div>
        {apply_section}
        <div style="margin-top: 32px; padding-top: 16px; border-top: 1px solid #e5e7eb;
                    font-size: 12px; color: #9ca3af;">
            This email was sent via Rekshift on behalf of your recruiter.
        </div>
    </div>
    """

    try:
        email_svc = EmailService()
        await email_svc.send_email(
            to=to,
            subject=payload.subject,
            html=html_content,
            text=payload.body,
        )
        logger.info(
            "[outreach_email] Sent to %s for job '%s'",
            to, payload.job_title or "—"
        )
        return ok({"sent": True, "to": to})

    except RuntimeError as e:
        # SMTP misconfigured
        return api_error(message=f"Email configuration error: {e}", status_code=503)
    except Exception as e:
        logger.error("[outreach_email] Failed: %s", e)
        return api_error(message=f"Failed to send email: {e}", status_code=500)


# ─── ATS Score Profile ────────────────────────────────────────────────────────

class AtsScoreProfileRequest(BaseModel):
    job_id: str
    profile: Dict[str, Any]


@router.post("/ats-score-profile")
async def ats_score_profile(
    payload: AtsScoreProfileRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    Run a detailed ATS analysis of a LinkedIn profile against a specific Job Description.
    Returns sub-scores for Skills, Experience, Education, and an overall weighted ATS score,
    plus strengths, gaps, and an AI recommendation.
    """
    db = get_db_admin_service()

    # Fetch the JD (direct sync call — correct pattern for this module)
    try:
        job_res = (
            db.client.from_("job_descriptions")
            .select("title, description, must_have_skills, good_to_have_skills, min_experience_years, role, level, location")
            .eq("id", payload.job_id)
            .eq("created_by", user.id)
            .maybe_single()
            .execute()
        )
        job = job_res.data if job_res else None
    except Exception as e:
        return api_error(message=f"Failed to fetch job: {e}", status_code=500)

    if not job:
        return api_error(message="Job not found.", status_code=404)

    p = payload.profile

    # Build compact profile summary for the prompt
    work = p.get("work_experience") or []
    edu = p.get("education") or []
    skills_raw = p.get("skills") or []
    skills_list = [s.get("name", str(s)) if isinstance(s, dict) else str(s) for s in skills_raw[:20]]

    work_summary = []
    for w in work[:5]:
        pos = w.get("position", "")
        company = w.get("company", "")
        start = w.get("start", "")
        end = w.get("end", "present")
        if pos or company:
            work_summary.append(f"{pos} at {company} ({start}–{end})".strip(" –"))

    edu_summary = []
    for e in edu[:3]:
        school = e.get("school", "")
        degree = e.get("degree", "")
        if school or degree:
            edu_summary.append(f"{degree} — {school}".strip(" —"))

    name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip()
    headline = p.get("headline", "")
    summary_text = p.get("summary", "")

    must_have = ", ".join(job.get("must_have_skills") or []) or "Not specified"
    good_to_have = ", ".join(job.get("good_to_have_skills") or []) or "None"
    min_exp = job.get("min_experience_years") or 0

    prompt = f"""You are a senior ATS (Applicant Tracking System) engine. Score this LinkedIn profile against the job description below.

=== JOB DESCRIPTION ===
Title: {job.get("title", "")}
Role Type: {job.get("role", "")} | Level: {job.get("level", "")}
Location: {job.get("location", "")}
Must-Have Skills: {must_have}
Nice-to-Have Skills: {good_to_have}
Min Experience: {min_exp} years
Description: {(job.get("description") or "")[:3000]}

=== CANDIDATE PROFILE ===
Name: {name}
Headline: {headline}
Summary: {(summary_text or "")[:1000]}
Skills: {", ".join(skills_list) or "Not listed"}
Work Experience:
{chr(10).join(f"  • {w}" for w in work_summary) or "  None listed"}
Education:
{chr(10).join(f"  • {e}" for e in edu_summary) or "  None listed"}

=== SCORING INSTRUCTIONS ===
Score each dimension from 0 to 100:

1. skills_score: How well do the candidate's skills match the MUST-HAVE and nice-to-have skills?
2. experience_score: Does the candidate's work history match the required experience level and domain?
3. education_score: Does their education match the role requirements?
4. overall_score: Weighted composite (skills 50%, experience 35%, education 15%)

Also provide:
- strengths: list of 2-4 key strong points (short phrases)
- gaps: list of 1-3 notable missing requirements (short phrases, or empty list if none)
- recommendation: one of "Highly Recommended", "Recommended", "Borderline", "Not Recommended"
- recommendation_reason: one sentence explaining the recommendation

Return ONLY valid JSON with no markdown:
{{
  "skills_score": 0,
  "experience_score": 0,
  "education_score": 0,
  "overall_score": 0,
  "strengths": ["..."],
  "gaps": ["..."],
  "recommendation": "...",
  "recommendation_reason": "..."
}}"""

    try:
        ai = get_ai()
        result = await ai.generate_json(prompt, temperature=0.1, max_tokens=600, timeout_s=30)
        if not isinstance(result, dict):
            raise ValueError(f"Invalid AI response type: {type(result)}")

        return ok({
            "skills_score": int(result.get("skills_score", 0)),
            "experience_score": int(result.get("experience_score", 0)),
            "education_score": int(result.get("education_score", 0)),
            "overall_score": int(result.get("overall_score", 0)),
            "strengths": result.get("strengths", []),
            "gaps": result.get("gaps", []),
            "recommendation": result.get("recommendation", ""),
            "recommendation_reason": result.get("recommendation_reason", ""),
        })
    except Exception as e:
        logger.error("[ats_score_profile] error: %s", e)
        return api_error(message=f"ATS scoring failed: {e}", status_code=502)


# ─── Add LinkedIn Profile as Candidate ───────────────────────────────────────

class AddAsCandidateRequest(BaseModel):
    job_id: str
    profile: Dict[str, Any]


@router.post("/add-as-candidate")
async def add_as_candidate(
    payload: AddAsCandidateRequest,
    user: ClerkUser = Depends(require_user),
):
    """
    Import a LinkedIn profile as a formal Candidate record linked to the given Job.

    Since LinkedIn doesn't expose email addresses, a synthetic placeholder email is generated:
        linkedin-{public_identifier}@linkedin-import.rekshift.com
    The recruiter can update the real email later from the Candidates page.
    """
    from datetime import datetime, timezone
    import uuid

    db = get_db_admin_service()

    # Verify job belongs to this recruiter
    try:
        job_res = (
            db.client.from_("job_descriptions")
            .select("id, title")
            .eq("id", payload.job_id)
            .eq("created_by", user.id)
            .maybe_single()
            .execute()
        )
        job = job_res.data if job_res else None
    except Exception as e:
        return api_error(message=f"Failed to verify job: {e}", status_code=500)

    if not job:
        return api_error(message="Job not found or access denied.", status_code=404)

    p = payload.profile
    first_name = (p.get("first_name") or "").strip()
    last_name = (p.get("last_name") or "").strip()
    full_name = f"{first_name} {last_name}".strip() or "LinkedIn Candidate"
    public_id = p.get("public_identifier") or p.get("provider_id") or str(uuid.uuid4())[:8]
    location = p.get("location", "")

    # Synthetic placeholder email (recruiter edits later)
    placeholder_email = f"linkedin-{public_id}@linkedin-import.rekshift.com"

    # Build structured resume_parsed_data from LinkedIn profile
    work = p.get("work_experience") or []
    edu = p.get("education") or []
    skills_raw = p.get("skills") or []
    skills_list = [s.get("name", str(s)) if isinstance(s, dict) else str(s) for s in skills_raw]

    resume_parsed_data = {
        "full_name": full_name,
        "email": None,
        "phone": None,
        "location": location,
        "headline": p.get("headline", ""),
        "summary": p.get("summary", ""),
        "skills": skills_list,
        "work_experience": work,
        "education": edu,
        "certifications": p.get("certifications") or [],
        "languages": p.get("languages") or [],
        "linkedin_url": (
            p.get("public_profile_url")
            or (f"https://www.linkedin.com/in/{public_id}" if public_id else "")
        ),
        "profile_picture_url": p.get("profile_picture_url"),
        "source": "linkedin_import",
        "linkedin_provider_id": p.get("provider_id"),
        "linkedin_match_score": p.get("match_score"),
        "linkedin_ai_summary": p.get("ai_summary"),
    }

    # Build plain-text resume from profile
    work_lines = []
    for w in work[:10]:
        pos = w.get("position", "")
        company = w.get("company", "")
        start = w.get("start", "")
        end = w.get("end", "Present")
        desc = w.get("description", "")
        line = f"{pos} at {company} ({start} – {end})"
        if desc:
            line += f"\n  {desc}"
        if pos or company:
            work_lines.append(line)

    edu_lines = []
    for e in edu[:5]:
        deg = e.get("degree", "")
        school = e.get("school", "")
        if deg or school:
            edu_lines.append(f"{deg} — {school}".strip(" —"))

    resume_text = f"""Name: {full_name}
LinkedIn: {resume_parsed_data['linkedin_url']}
Location: {location}
Headline: {p.get('headline', '')}

Summary:
{p.get('summary', 'N/A')}

Skills:
{', '.join(skills_list) or 'N/A'}

Work Experience:
{chr(10).join(work_lines) or 'N/A'}

Education:
{chr(10).join(edu_lines) or 'N/A'}
"""

    now = datetime.now(timezone.utc).isoformat()

    # Check if a candidate with this placeholder email already exists
    try:
        existing_res = (
            db.client.from_("candidates")
            .select("id")
            .eq("email", placeholder_email)
            .limit(1)
            .execute()
        )
        existing = existing_res.data if existing_res else []
    except Exception:
        existing = []

    if existing and isinstance(existing, list) and existing[0].get("id"):
        candidate_id = existing[0]["id"]
        already_exists = True
    else:
        new_cand_id = str(uuid.uuid4())
        insert_row = {
            "id": new_cand_id,
            "full_name": full_name,
            "email": placeholder_email,
            "location": location,
            "resume_parsed_data": resume_parsed_data,
            "resume_text": resume_text,
        }

        try:
            ins_res = db.client.from_("candidates").insert(insert_row).execute()
            ins_data = ins_res.data if ins_res else []
        except Exception as e:
            return api_error(message=f"Failed to create candidate record: {e}", status_code=500)

        if not ins_data:
            return api_error(message="Failed to create candidate record.", status_code=500)
        candidate_id = ins_data[0]["id"]
        already_exists = False

    # Check if already linked to this job
    try:
        app_res = (
            db.client.from_("job_applications")
            .select("id")
            .eq("candidate_id", candidate_id)
            .eq("job_id", payload.job_id)
            .maybe_single()
            .execute()
        )
        existing_app = app_res.data if app_res else None
    except Exception:
        existing_app = None

    if isinstance(existing_app, dict) and existing_app.get("id"):
        return ok({
            "candidate_id": candidate_id,
            "job_application_id": existing_app["id"],
            "already_exists": True,
            "message": f"{full_name} is already in your candidates for this job.",
        })

    # Link to job
    try:
        app_ins_res = (
            db.client.from_("job_applications")
            .insert({
                "candidate_id": candidate_id,
                "job_id": payload.job_id,
                "status": "applied",
                "applied_at": now,
                "candidate_overrides": {
                    "full_name": full_name,
                    "source": "linkedin_import",
                    "linkedin_url": resume_parsed_data["linkedin_url"],
                    "profile_picture_url": p.get("profile_picture_url"),
                },
            })
            .execute()
        )
        app_data = app_ins_res.data if app_ins_res else []
    except Exception as e:
        return api_error(message=f"Failed to link candidate to job: {e}", status_code=500)

    app_id = app_data[0]["id"] if isinstance(app_data, list) and app_data else None

    logger.info(
        "[add_as_candidate] LinkedIn profile '%s' added to job '%s' (candidate_id=%s)",
        full_name, payload.job_id, candidate_id,
    )

    return ok({
        "candidate_id": candidate_id,
        "job_application_id": app_id,
        "already_exists": already_exists,
        "message": f"{full_name} has been added to your candidates.",
        "placeholder_email": placeholder_email,
    })

